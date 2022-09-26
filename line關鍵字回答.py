# 技術文章
# https://developers.line.biz/en/docs/messaging-api/sending-messages/#methods-of-sending-message

# 一定要改的地方～～
# 請打開 ine.xlsx 中的  setup 改 auth_token 	接收人的id
# auth_token="2bSr7EUf1Ip1OVYBgh7HgCpdrPj/KyphlQQpQKkyUN3NVuZlQ36dp2b3O6imlhLk2LunCyORmlfQHb9tNXaul716CTlDNSD78PPUaL3ZM51a5yvLf6oWBQ4zyeBcAgD8Agc478HreI5bD5ArdslHrwdB04t89/1O/w1cDnyilFU="
#接收人的id=["U22869b7bf55a2026578867d615fe8c11"]

import requests
import json
from openpyxl import Workbook     # pip install openpyxl
import time                       # 時間
from openpyxl import load_workbook
from sys import version as python_version
from cgi import parse_header, parse_multipart
import socketserver as socketserver
import http.server
from http.server import SimpleHTTPRequestHandler as RequestHandler
from urllib.parse import parse_qs

"""
技術文件
https://developers.line.biz/en/reference/messaging-api/#set-webhook-endpoint-url


curl -X PUT \
-H 'Authorization: Bearer {CHANNEL_ACCESS_TOKEN}' \
-H 'Content-Type:application/json' \
-d '{"endpoint":"https://example.com/hoge"}' \
https://api.line.me/v2/bot/channel/webhook/endpoint
"""


#自動更新ngrok
import requests
import json
from openpyxl import Workbook     # pip install openpyxl
import time                       # 時間
from openpyxl import load_workbook
from sys import version as python_version
from cgi import parse_header, parse_multipart
import socketserver as socketserver
import http.server
from http.server import SimpleHTTPRequestHandler as RequestHandler
from urllib.parse import parse_qs
import PKLineAP


#  SSL  處理，  https    SSSSSS 就需要加上以下2行
import ssl
ssl._create_default_https_context = ssl._create_unverified_context    # 因.urlopen發生問題，將ssl憑證排除


ngrokHTTP=PKLineAP.ngrok_啟動()
auth_token,接收人的id=PKLineAP.Line_讀取設定檔Excel('line.xlsx')
t1=PKLineAP.Line_設定Webhook(ngrokHTTP,auth_token)
print(t1)

wb = load_workbook('line.xlsx')  # 讀取檔案
# 方法一打開第一個 工作表單
sheetSetup = wb["setup"]         # 打開一個工作欄
auth_token=sheetSetup.cell(row=2, column=1).value
# 接收人的id=sheetSetup.cell(row=2, column=2).value

m_url=ngrokHTTP

sheet問答題 = wb["問答題"]         # 打開一個工作欄



def openpyxl_GetRow(sheet,用戶輸入的文字="",關鍵字的col=1,回傳答案欄位=2):
    row1=1
    回答=""
    while row1<=sheet.max_row:
        關鍵字Value=sheet.cell(row=row1, column=關鍵字的col).value  # 取得資料
        關鍵字Value=str(關鍵字Value)
        if(用戶輸入的文字.find(關鍵字Value)>=0):                          # -1 查無此資料
            回答=sheet.cell(row=row1, column=回傳答案欄位).value    # 取得資料
            break                                                # 結束迴圈
        row1=row1+1
    return 回答

def Line_回傳Text(str):
    回傳值 = [
        {
            "type": "text",
            "text":str
        }
    ]

    return 回傳值

def Line_訂位(str):
    回傳值 = [
        {
            "type": "text",
            "text": "您輸入的是:" + str,
            "text": "請洽詢電話或線上訂位:"

        },
        {
            "type": "sticker",
            "packageId": "789",
            "stickerId": "10857"
        }

    ]

    return 回傳值


def Line_回傳地址(str):
    回傳值 = [
        {
            "type": "text",
            "text": "桃園市中壢區環中東路"
        },
        {
            "type": "location",
            "title": "桃園市中壢區",
            "address": "桃園市中壢區環中東路",
            "latitude": 24.9591722,
            "longitude": 121.248982
        },
        {
            "type": "image",
            "originalContentUrl": m_url + "/1.jpg",
            "previewImageUrl": m_url + "/1.jpg"
        },

        {
            "type": "video",
            "originalContentUrl": m_url+ "/seapasta.mp4",
            "previewImageUrl": m_url + "/1.jpg",
            "trackingId": "track-id"
        },
        {
            "type": "audio",
            "originalContentUrl": m_url + "/1.mp3",
            "duration": 105000  # 60000
        }
    ]

    return 回傳值
def Line_其他回答(str):
    回傳值 = [
        {
            "type": "text",
            "text": "請重新輸入關鍵用語，謝謝。"
        },

    ]

    return 回傳值

def Line_處理用的問題(用戶輸入的文字):
    回答 =openpyxl_GetRow(sheet問答題,用戶輸入的文字,關鍵字的col=1,回傳答案欄位=2)
    if(回答!=""):
        回傳值 =Line_回傳Text(回答)
    elif(用戶輸入的文字=="訂位"):
        回傳值 =Line_訂位(回答)
    elif (用戶輸入的文字 == "地址"):
        回傳值 = Line_回傳地址(回答)
    else:
        回傳值 =Line_其他回答(回答)
    return 回傳值

#line 內容
class MyHandler(RequestHandler):
    def do_POST(self):
        varLen = int(self.headers['Content-Length'])        # 取得讀取進來的網路資料長度
        if varLen > 0:
            post_data = self.rfile.read(varLen)             # 讀取傳過來的資料
            data = json.loads(post_data)                    # 把字串 轉成JSON
            print(data)
            replyToken=data['events'][0]['replyToken']       # 回傳要用Token
            userId3=data['events'][0]['source']['userId']    # 傳資料過來的使用者是誰
            用戶輸入的文字=data['events'][0]['message']['text']        # 用戶的傳遞過來的文字內容
            傳過來的資料型態=data['events'][0]['message']['type'] # 傳過來的資料型態

        回傳值=Line_處理用的問題(用戶輸入的文字)

        message = {
            "replyToken": replyToken,
            "messages": 回傳值
        }

        # 資料回傳 到 Line 的 https 伺服器
        hed = {'Authorization': 'Bearer ' + auth_token}
        url = 'https://api.line.me/v2/bot/message/reply'
        self.send_response(200)
        self.end_headers()
        requests.post(url, json=message, headers=hed)      # 把資料HTTP POST送出去




socketserver.TCPServer.allow_reuse_address = True              # 可以重複使用IP
httpd = socketserver.TCPServer(('0.0.0.0', 8888), MyHandler)  # 啟動WebServer   :8888
try:
    print("伺服器啟動 "+ngrokHTTP+":8888")
    httpd.serve_forever()                          # 等待用戶使用 WebServer
except:
    print("Closing the server.")
    httpd.server_close()                           # 關閉 WebServer
    print(t1)
    #PKLineAP.ngrok_持續執行()
    PKLineAP.ngrok_關閉()
    print("Server closed.")



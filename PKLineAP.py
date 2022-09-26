import json

import requests
from openpyxl import load_workbook

# "line設定.json" 一定要改的地方～～
auth_token=""
YouruserID=""
userId=""
groupID=""


from subprocess import Popen, PIPE
import os
import socketserver as socketserver
import http.server
from http.server import SimpleHTTPRequestHandler as RequestHandler
from urllib.parse import urlparse
import json

import subprocess
import urllib.parse
import os

def String_HTML轉print(str1):
    str1=str1.replace("<br>","\n")
    str1=str1.replace("<hr>","\n\n")
    return str1


def 外掛_桃園ubike(sna):
    exe2 = "python3 youbike.py "+sna
    str1 = 執行外部程式(exe2)
    return String_HTML轉print(str1)

def 執行外部程式(exe2):
    html1=""
    #exe2 = "python youbike.py 公園"
    with Popen(exe2,stdout=PIPE, stderr=PIPE, shell=True) as p:
        output, errors = p.communicate()
        list1 = output.decode('utf-8').splitlines()
        for str1 in list1:
            html1 = html1 + str1
    return html1
# auth_token,接收人的id=Line_讀取設定檔Excel('line.xlsx')
def Line_讀取設定檔Excel(filename):
    global auth_token,YouruserID,userId,groupID,接收人的id
    wb = load_workbook(filename) # 'line.xlsx')  # 讀取檔案
    # 方法一打開第一個 工作表單
    sheetSetup = wb["setup"]         # 打開一個工作欄
    auth_token=sheetSetup.cell(row=2, column=1).value
    接收人的id=sheetSetup.cell(row=2, column=2).value
    return auth_token,接收人的id


def Line_設定Webhook(ngrokHTTP,auth_token):
    # 資料回傳 到 Line 的 https 伺服器
    # ngrokHTTP = "https://d5ce-114-44-20-217.ngrok.io "
    message = {"endpoint": ngrokHTTP}
    hed = {'Authorization': 'Bearer ' + auth_token}
    url = 'https://api.line.me/v2/bot/channel/webhook/endpoint'
    t1 = requests.put(url, json=message, headers=hed)  # 把資料HTTP POST送出去
    print(t1)
    return t1

# https://jcshawn.com/linebot-user-profile/
# https://developers.line.biz/en/reference/line-login-v2/#get-user-profile
"""
curl -v -X GET https://api.line.me/v2/profile \
-H 'Authorization: Bearer {access token}'
"""
def Line_取得用戶的資訊(userId,auth_token):
    # 資料回傳 到 Line 的 https 伺服器
    # ngrokHTTP = "https://d5ce-114-44-20-217.ngrok.io "

    hed = {'Authorization': 'Bearer ' + auth_token}
    url = 'https://api.line.me/v2/bot/profile/'+userId
    t1 = requests.get(url, headers=hed)  # 把資料HTTP POST送出去
    print(t1)
    if t1.status_code == 200:
        return  t1.text
    return t1





def Line_讀取設定檔(filename):
    global auth_token,YouruserID,userId,groupID
    # read a json file and return a dictionary
    with open(filename, 'r', encoding='utf-8') as f:
       dict1 = json.load(f)
       auth_token = dict1.get('auth_token',"")
       YouruserID = dict1.get('YouruserID',"")
       userId = dict1.get('userId',"")
       groupID = dict1.get('groupID',"")




"""
參考資料
https://pyngrok.readthedocs.io/en/latest/
"""
from pyngrok import ngrok     # pip install pyngrok

def ngrok_啟動(port=8888,protocol="http"):
    # Open a HTTP tunnel on the default port 80
    # <NgrokTunnel: "http://<public_sub>.ngrok.io" -> "http://localhost:80">
    # http_tunnel = ngrok.connect()
    # Open a SSH tunnel
    # <NgrokTunnel: "tcp://0.tcp.ngrok.io:12345" -> "localhost:8888">

    ngrok_關閉()
    ssh_tunnel = ngrok.connect(port,protocol)  # 8888, "http")        # 開啟 localhost:8888 的 HTTP 連線
    publicIP=""
    tunnels = ngrok.get_tunnels()                   # 取得所有的tunnel
    for x in tunnels:
        if x.config["addr"].lower()=="http://localhost:"+str(port):
            if x.proto.lower() == "https":
                publicIP=x.data["public_url"]
                print(publicIP)         # 取得第二個tunnel的public_url
                break
    return publicIP


def ngrok_持續執行():
    ngrok_process = ngrok.get_ngrok_process()
    try:
        # Block until CTRL-C or some other terminating event
        print(" CTRL-C 關閉程式")
        ngrok_process.proc.wait()
    except KeyboardInterrupt:
        ngrok_關閉()


def ngrok_關閉():
    print(" Shutting down server.")
    tunnels = ngrok.get_tunnels()                   # 取得所有的tunnel
    for x in tunnels:
        ngrok.disconnect(x.public_url)  # 取消連線
    ngrok.kill()






import uuid
def UUID_產生器():
    return  str(uuid.uuid4())


def String_split(str1,len=5000):
    list1=[]
    for i in range(0,len(str1),len):
        list1.append(str1[i:i+len])
    return list1


def Line_回送文字(replyToken="",text="",userId=""):
    if text=="":
        str1="你的 User Id: " + userId + "\n 傳過來的文字 Text:" + text
    else:
        str1=text

    message = {
        "replyToken": replyToken,
        "messages": [
            {
                "type": "text",
                "text": str1
            }
        ]
    }
    return Line_回送(message)


def Line_回送(message=""):

    hed = {'Authorization': 'Bearer ' + auth_token}
    url = 'https://api.line.me/v2/bot/message/reply'
    response = requests.post(url, json=message, headers=hed)
    return response


def Line_廣播推送(toID="",str1="hello"):
    message={
            "to": toID,
            "messages":[
                {
                    "type":"text",
                    "text":str1
                }
            ]
        }


    UUID=UUID_產生器()
    hed = {'Content-Type': 'application/json',
           'Authorization': 'Bearer ' + auth_token,
           'X-Line-Retry-Key':UUID}
    url = 'https://api.line.me/v2/bot/message/push'
    response = requests.post(url, json=message, headers=hed)
    print(response)

"""
curl -v -X POST https://api.line.me/v2/bot/message/broadcast \
-H 'Content-Type: application/json' \
-H 'Authorization: Bearer {channel access token}' \
-H 'X-Line-Retry-Key: {UUID}' \
-d '{
    "messages":[
        {
            "type":"text",
            "text":"Hello, world1"
        },
        {
            "type":"text",
            "text":"Hello, world2"
        }
    ]
}'
"""

def Line_廣播到所有用戶(str1="hello"):
    message={
            "messages":[
                {
                    "type":"text",
                    "text":str1
                }
            ]
        }


    UUID=UUID_產生器()
    hed = {'Content-Type': 'application/json',
           'Authorization': 'Bearer ' + auth_token,
           'X-Line-Retry-Key':UUID}
    url = 'https://api.line.me/v2/bot/message/broadcast'
    response = requests.post(url, json=message, headers=hed)
    print(response)